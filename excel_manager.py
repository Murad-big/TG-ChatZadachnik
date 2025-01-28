import openpyxl
from config import EXCEL_FILE
import os


def read_excel_data():
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook["Workers"]

    names, desc, dates, statuses = [], [], [], []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if all(cell is None for cell in row):
            continue
        names.append(row[0])
        desc.append(row[1])
        dates.append(row[2])
        statuses.append(row[3])

    return names, desc, dates, statuses

def write_excel_data(names, desc, dates, statuses):
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook["Workers"]

    for row_idx in range(2, sheet.max_row + 1):
        for col_idx in range(1, 5):
            sheet.cell(row=row_idx, column=col_idx).value = None

    for i in range(len(names)):
        sheet.cell(row=i + 2, column=1, value=names[i])
        sheet.cell(row=i + 2, column=2, value=desc[i])
        sheet.cell(row=i + 2, column=3, value=dates[i])
        sheet.cell(row=i + 2, column=4, value=statuses[i])

    workbook.save(EXCEL_FILE)

def initialize_excel():
    if not os.path.exists(EXCEL_FILE):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Workers"
        sheet.append(["Username", "Task", "Deadline", "Status"])
        workbook.save(EXCEL_FILE)

def add_task_to_excel(username, task, deadline):
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook["Workers"]
    sheet.append([username, task, deadline, "Незакончен"])
    workbook.save(EXCEL_FILE)